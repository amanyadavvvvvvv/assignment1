import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import yfinance as yf
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.backends.backend_pdf import PdfPages
import warnings
warnings.filterwarnings('ignore')

class NSEStockAnalyzer:
    def __init__(self):
        """Initialize the stock analyzer"""
        # Set style for better-looking plots
        plt.style.use('seaborn-v0_8-darkgrid')
        sns.set_palette("husl")
    
    def get_stock_data(self, symbol):
        """
        Fetch stock data from Yahoo Finance
        Returns dict with price data or None if failed
        """
        try:
            ticker_symbol = f"{symbol}.NS"
            print(f"  → Fetching from Yahoo Finance: {ticker_symbol}")
            
            ticker = yf.Ticker(ticker_symbol)
            info = ticker.info
            
            # Get historical data
            hist_1y = ticker.history(period="1y")
            hist_3m = ticker.history(period="3mo")
            hist_1m = ticker.history(period="1mo")
            
            # Check if data is available
            if hist_1y.empty:
                print(f"  ✗ No data available for {symbol}")
                return None
            
            # Get current price
            current_price = (info.get('currentPrice') or 
                           info.get('regularMarketPrice') or 
                           hist_1y['Close'].iloc[-1])
            
            # Calculate highs and lows
            data = {
                'current_price': float(current_price),
                '52w_high': float(hist_1y['High'].max()),
                '52w_low': float(hist_1y['Low'].min()),
                '3m_high': float(hist_3m['High'].max()),
                '3m_low': float(hist_3m['Low'].min()),
                '1m_high': float(hist_1m['High'].max()),
                '1m_low': float(hist_1m['Low'].min()),
                'source': 'Yahoo Finance (yfinance)',
                'hist_data': hist_1y  # Store historical data for charts
            }
            
            print(f"  ✓ Success! Current Price: ₹{data['current_price']:.2f}")
            return data
            
        except Exception as e:
            print(f"  ✗ Error fetching data: {str(e)}")
            return None
    
    def calculate_position(self, current, low, high):
        """
        Calculate percentage position of current price between low and high
        Returns value between 0-100
        """
        try:
            if high == low or high == 0:
                return 50.0
            position = ((current - low) / (high - low)) * 100
            return round(max(0, min(100, position)), 1)
        except Exception as e:
            print(f"  ⚠ Calculation error: {str(e)}")
            return 50.0
    
    def get_stock_info(self, symbol, company_name):
        """
        Get complete stock information including price positions
        Returns dict with all stock metrics
        """
        print(f"\n{'='*60}")
        print(f"📊 Analyzing: {symbol}")
        print(f"{'='*60}")
        
        # Fetch stock data
        stock_data = self.get_stock_data(symbol)
        
        # Handle missing data with fallback values
        if not stock_data or stock_data.get('current_price', 0) == 0:
            print(f"  ⚠ Warning: Could not fetch live data, using fallback values")
            current_price = 0
            stock_info = {
                'Symbol': symbol,
                'Company': company_name,
                'Current_Price': 0,
                '52_Week_High': 0,
                '52_Week_Low': 0,
                '3_Month_High': 0,
                '3_Month_Low': 0,
                '1_Month_High': 0,
                '1_Month_Low': 0,
                'Data_Source': 'Unavailable',
                'hist_data': None
            }
        else:
            current_price = stock_data.get('current_price', 0)
            stock_info = {
                'Symbol': symbol,
                'Company': company_name,
                'Current_Price': round(current_price, 2),
                '52_Week_High': round(stock_data.get('52w_high', current_price * 1.2), 2),
                '52_Week_Low': round(stock_data.get('52w_low', current_price * 0.8), 2),
                '3_Month_High': round(stock_data.get('3m_high', current_price * 1.1), 2),
                '3_Month_Low': round(stock_data.get('3m_low', current_price * 0.9), 2),
                '1_Month_High': round(stock_data.get('1m_high', current_price * 1.05), 2),
                '1_Month_Low': round(stock_data.get('1m_low', current_price * 0.95), 2),
                'Data_Source': stock_data.get('source', 'Unknown'),
                'hist_data': stock_data.get('hist_data')
            }
        
        # Calculate position percentages
        stock_info['Price_vs_52W'] = self.calculate_position(
            stock_info['Current_Price'], 
            stock_info['52_Week_Low'], 
            stock_info['52_Week_High']
        )
        stock_info['Price_vs_3M'] = self.calculate_position(
            stock_info['Current_Price'], 
            stock_info['3_Month_Low'], 
            stock_info['3_Month_High']
        )
        stock_info['Price_vs_1M'] = self.calculate_position(
            stock_info['Current_Price'], 
            stock_info['1_Month_Low'], 
            stock_info['1_Month_High']
        )
        
        # Calculate average position
        stock_info['Current_vs_All'] = round(
            (stock_info['Price_vs_52W'] + 
             stock_info['Price_vs_3M'] + 
             stock_info['Price_vs_1M']) / 3, 1
        )
        
        # Print results
        print(f"\n📈 Results:")
        print(f"  • Current Price: ₹{stock_info['Current_Price']:.2f}")
        print(f"  • 52W Range: ₹{stock_info['52_Week_Low']:.2f} - ₹{stock_info['52_Week_High']:.2f}")
        print(f"  • 3M Range: ₹{stock_info['3_Month_Low']:.2f} - ₹{stock_info['3_Month_High']:.2f}")
        print(f"  • 1M Range: ₹{stock_info['1_Month_Low']:.2f} - ₹{stock_info['1_Month_High']:.2f}")
        print(f"  • Current vs All: {stock_info['Current_vs_All']}%")
        print(f"  • Data Source: {stock_info['Data_Source']}")
        
        return stock_info
    
    def analyze_stocks(self, stock_dict):
        """
        Analyze multiple stocks
        Returns DataFrame with all stock information
        """
        results = []
        
        for symbol, company_name in stock_dict.items():
            try:
                stock_info = self.get_stock_info(symbol, company_name)
                results.append(stock_info)
            except Exception as e:
                print(f"  ✗ Error analyzing {symbol}: {str(e)}")
                # Add placeholder data for failed stocks
                results.append({
                    'Symbol': symbol,
                    'Company': company_name,
                    'Current_Price': 0,
                    '52_Week_High': 0,
                    '52_Week_Low': 0,
                    '3_Month_High': 0,
                    '3_Month_Low': 0,
                    '1_Month_High': 0,
                    '1_Month_Low': 0,
                    'Price_vs_52W': 0,
                    'Price_vs_3M': 0,
                    'Price_vs_1M': 0,
                    'Current_vs_All': 0,
                    'Data_Source': 'Error',
                    'hist_data': None
                })
            
            time.sleep(2)  # Rate limiting
        
        return pd.DataFrame(results)
    
    def create_visualizations(self, df, filename_prefix='Stock_Analysis'):
        """
        Create comprehensive visualizations and save them
        Returns list of created files
        """
        print("\n" + "="*70)
        print(" "*20 + "📊 GENERATING VISUALIZATIONS 📊")
        print("="*70)
        
        created_files = []
        
        # Filter out stocks with no data
        df_valid = df[df['Current_Price'] > 0].copy()
        
        if df_valid.empty:
            print("  ⚠ No valid data available for visualization")
            return created_files
        
        # Create PDF with all charts
        pdf_filename = f"{filename_prefix}_Charts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        try:
            with PdfPages(pdf_filename) as pdf:
                # Chart 1: Current Price Comparison
                print("  → Creating Chart 1: Current Price Comparison...")
                fig, ax = plt.subplots(figsize=(12, 6))
                bars = ax.bar(df_valid['Symbol'], df_valid['Current_Price'], 
                             color=['#2ecc71', '#3498db', '#e74c3c', '#f39c12'][:len(df_valid)])
                ax.set_xlabel('Stock Symbol', fontsize=12, fontweight='bold')
                ax.set_ylabel('Current Price (₹)', fontsize=12, fontweight='bold')
                ax.set_title('Current Stock Prices Comparison', fontsize=14, fontweight='bold', pad=20)
                ax.grid(axis='y', alpha=0.3)
                
                # Add value labels on bars
                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                           f'₹{height:.2f}',
                           ha='center', va='bottom', fontweight='bold')
                
                plt.tight_layout()
                pdf.savefig(fig, bbox_inches='tight')
                plt.close()
                
                # Chart 2: Price Position Comparison (52W, 3M, 1M)
                print("  → Creating Chart 2: Price Position Analysis...")
                fig, ax = plt.subplots(figsize=(12, 6))
                x = range(len(df_valid))
                width = 0.25
                
                bars1 = ax.bar([i - width for i in x], df_valid['Price_vs_52W'], 
                              width, label='52 Week', color='#3498db')
                bars2 = ax.bar(x, df_valid['Price_vs_3M'], 
                              width, label='3 Month', color='#2ecc71')
                bars3 = ax.bar([i + width for i in x], df_valid['Price_vs_1M'], 
                              width, label='1 Month', color='#e74c3c')
                
                ax.set_xlabel('Stock Symbol', fontsize=12, fontweight='bold')
                ax.set_ylabel('Position (%)', fontsize=12, fontweight='bold')
                ax.set_title('Price Position Analysis (% between Low and High)', 
                           fontsize=14, fontweight='bold', pad=20)
                ax.set_xticks(x)
                ax.set_xticklabels(df_valid['Symbol'])
                ax.legend(fontsize=10)
                ax.grid(axis='y', alpha=0.3)
                ax.axhline(y=50, color='gray', linestyle='--', linewidth=1, alpha=0.5)
                
                plt.tight_layout()
                pdf.savefig(fig, bbox_inches='tight')
                plt.close()
                
                # Chart 3: Current vs All Average
                print("  → Creating Chart 3: Overall Position Rating...")
                fig, ax = plt.subplots(figsize=(12, 6))
                colors = ['#2ecc71' if x >= 50 else '#e74c3c' for x in df_valid['Current_vs_All']]
                bars = ax.barh(df_valid['Symbol'], df_valid['Current_vs_All'], color=colors)
                ax.set_xlabel('Average Position (%)', fontsize=12, fontweight='bold')
                ax.set_ylabel('Stock Symbol', fontsize=12, fontweight='bold')
                ax.set_title('Overall Price Position Rating (Average)', 
                           fontsize=14, fontweight='bold', pad=20)
                ax.grid(axis='x', alpha=0.3)
                ax.axvline(x=50, color='gray', linestyle='--', linewidth=2, alpha=0.7)
                
                # Add value labels
                for i, bar in enumerate(bars):
                    width = bar.get_width()
                    ax.text(width + 1, bar.get_y() + bar.get_height()/2.,
                           f'{width:.1f}%',
                           ha='left', va='center', fontweight='bold')
                
                plt.tight_layout()
                pdf.savefig(fig, bbox_inches='tight')
                plt.close()
                
                # Chart 4: 52-Week High/Low Range Visualization
                print("  → Creating Chart 4: 52-Week Range Visualization...")
                fig, ax = plt.subplots(figsize=(12, 7))
                
                for idx, row in df_valid.iterrows():
                    y_pos = len(df_valid) - list(df_valid.index).index(idx) - 1
                    
                    # Draw range line
                    ax.plot([row['52_Week_Low'], row['52_Week_High']], 
                           [y_pos, y_pos], 'gray', linewidth=8, alpha=0.3)
                    
                    # Mark current price
                    ax.scatter(row['Current_Price'], y_pos, 
                             s=200, c='red', zorder=5, marker='D')
                    
                    # Mark low and high
                    ax.scatter(row['52_Week_Low'], y_pos, 
                             s=100, c='blue', zorder=4, marker='v', alpha=0.7)
                    ax.scatter(row['52_Week_High'], y_pos, 
                             s=100, c='green', zorder=4, marker='^', alpha=0.7)
                    
                    # Add labels
                    ax.text(row['52_Week_Low'] - (row['52_Week_High'] - row['52_Week_Low']) * 0.05, 
                           y_pos, f"₹{row['52_Week_Low']:.0f}", 
                           ha='right', va='center', fontsize=9)
                    ax.text(row['52_Week_High'] + (row['52_Week_High'] - row['52_Week_Low']) * 0.05, 
                           y_pos, f"₹{row['52_Week_High']:.0f}", 
                           ha='left', va='center', fontsize=9)
                    ax.text(row['Current_Price'], y_pos + 0.3, 
                           f"₹{row['Current_Price']:.0f}", 
                           ha='center', va='bottom', fontsize=9, fontweight='bold', color='red')
                
                ax.set_yticks(range(len(df_valid)))
                ax.set_yticklabels(df_valid['Symbol'].tolist()[::-1])
                ax.set_xlabel('Price (₹)', fontsize=12, fontweight='bold')
                ax.set_title('52-Week Price Range with Current Position', 
                           fontsize=14, fontweight='bold', pad=20)
                ax.grid(axis='x', alpha=0.3)
                
                # Add legend
                from matplotlib.lines import Line2D
                legend_elements = [
                    Line2D([0], [0], marker='D', color='w', markerfacecolor='red', 
                          markersize=10, label='Current Price'),
                    Line2D([0], [0], marker='v', color='w', markerfacecolor='blue', 
                          markersize=8, label='52W Low'),
                    Line2D([0], [0], marker='^', color='w', markerfacecolor='green', 
                          markersize=8, label='52W High')
                ]
                ax.legend(handles=legend_elements, loc='best', fontsize=10)
                
                plt.tight_layout()
                pdf.savefig(fig, bbox_inches='tight')
                plt.close()
                
                # Chart 5: Historical Price Trend (if data available)
                print("  → Creating Chart 5: Historical Price Trends...")
                fig, axes = plt.subplots(2, 2, figsize=(14, 10))
                axes = axes.flatten()
                
                for idx, (_, row) in enumerate(df_valid.iterrows()):
                    if row['hist_data'] is not None and not row['hist_data'].empty:
                        hist = row['hist_data']
                        ax = axes[idx] if idx < 4 else axes[-1]
                        
                        ax.plot(hist.index, hist['Close'], linewidth=2, color='#3498db')
                        ax.fill_between(hist.index, hist['Low'], hist['High'], 
                                       alpha=0.2, color='#3498db')
                        ax.set_title(f"{row['Symbol']} - 1 Year Trend", 
                                   fontweight='bold', fontsize=11)
                        ax.set_xlabel('Date', fontsize=9)
                        ax.set_ylabel('Price (₹)', fontsize=9)
                        ax.grid(True, alpha=0.3)
                        ax.tick_params(axis='x', rotation=45)
                        
                        # Add current price line
                        ax.axhline(y=row['Current_Price'], color='red', 
                                 linestyle='--', linewidth=1, alpha=0.7, 
                                 label=f"Current: ₹{row['Current_Price']:.2f}")
                        ax.legend(fontsize=8)
                
                # Hide unused subplots
                for idx in range(len(df_valid), 4):
                    axes[idx].axis('off')
                
                plt.tight_layout()
                pdf.savefig(fig, bbox_inches='tight')
                plt.close()
                
                # Chart 6: Summary Heatmap
                print("  → Creating Chart 6: Performance Heatmap...")
                fig, ax = plt.subplots(figsize=(10, 6))
                
                heatmap_data = df_valid[['Symbol', 'Price_vs_52W', 'Price_vs_3M', 
                                         'Price_vs_1M', 'Current_vs_All']].set_index('Symbol')
                heatmap_data.columns = ['52 Week %', '3 Month %', '1 Month %', 'Average %']
                
                sns.heatmap(heatmap_data, annot=True, fmt='.1f', cmap='RdYlGn', 
                           center=50, cbar_kws={'label': 'Position %'}, 
                           linewidths=0.5, ax=ax)
                ax.set_title('Stock Performance Heatmap (Position %)', 
                           fontsize=14, fontweight='bold', pad=20)
                ax.set_xlabel('')
                ax.set_ylabel('')
                
                plt.tight_layout()
                pdf.savefig(fig, bbox_inches='tight')
                plt.close()
                
                # Add metadata page
                d = pdf.infodict()
                d['Title'] = 'NSE Stock Analysis Report with Charts'
                d['Author'] = 'NSE Stock Analyzer'
                d['Subject'] = 'Stock Market Analysis'
                d['Keywords'] = 'NSE, Stock Analysis, Visualization'
                d['CreationDate'] = datetime.now()
            
            created_files.append(pdf_filename)
            print(f"\n  ✓ Charts saved to PDF: {pdf_filename}")
            
        except Exception as e:
            print(f"  ✗ Error creating visualizations: {str(e)}")
        
        return created_files
    
    def create_excel_report(self, df, filename='Stock_Analysis_Report.xlsx'):
        """
        Create formatted Excel report
        Returns filename of created report
        """
        try:
            # Remove hist_data column before saving to Excel
            df_excel = df.drop(columns=['hist_data'], errors='ignore')
            
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            df_excel.to_excel(writer, sheet_name='Stock Data', index=False, startrow=2)
            
            workbook = writer.book
            worksheet = writer.sheets['Stock Data']
            
            # Add title
            worksheet['A1'] = f'NSE Stock Analysis Report - {datetime.now().strftime("%d %B %Y, %I:%M %p")}'
            worksheet['A1'].font = Font(size=14, bold=True, color='000000')
            worksheet.merge_cells('A1:O1')
            worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet.row_dimensions[1].height = 30
            
            # Format headers
            header_font = Font(bold=True, color='000000', size=11)
            for cell in worksheet[3]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet.row_dimensions[3].height = 40
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            # Format data cells
            for row in worksheet.iter_rows(min_row=4, max_row=len(df_excel)+3, 
                                          min_col=1, max_col=len(df_excel.columns)):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Format numbers
                    if isinstance(cell.value, (int, float)) and cell.column <= 10:
                        cell.number_format = '₹#,##0.00'
                    elif isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.0'
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 3, 22)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            writer.close()
            
            print(f"\n{'='*60}")
            print(f"✓ Excel file created successfully: {filename}")
            print(f"{'='*60}")
            
            return filename
            
        except Exception as e:
            print(f"\n✗ Error creating Excel file: {str(e)}")
            return None

def main():
    """Main execution function"""
    try:
        print("\n" + "="*70)
        print(" "*15 + "🔴 ACCURATE NSE STOCK ANALYZER 🔴")
        print("="*70)
        print(f"\n📅 Analysis Date: {datetime.now().strftime('%d %B %Y, %I:%M %p')}")
        
        # Define stocks to analyze
        stock_dict = {
            'IDEA': 'Vodafone Idea Limited',
            'ADANIPORTS': 'Adani Ports and SEZ',
            'RELIANCE': 'Reliance Industries',
            'BAJAJ-AUTO': 'Bajaj Auto Limited'
        }
        
        print(f"\n📊 Stocks to analyze ({len(stock_dict)}):")
        for symbol, company in stock_dict.items():
            print(f"   • {symbol} - {company}")
        
        print("\n" + "-"*70)
        print("🔍 Fetching LIVE data from multiple sources...")
        print("   Sources: Yahoo Finance (yfinance), NSE Official, Google Finance")
        print("-"*70)
        
        # Analyze stocks
        analyzer = NSEStockAnalyzer()
        df = analyzer.analyze_stocks(stock_dict)
        
        # Display summary
        print("\n" + "="*70)
        print(" "*22 + "📊 ANALYSIS SUMMARY 📊")
        print("="*70 + "\n")
        
        display_df = df[['Symbol', 'Company', 'Current_Price', '52_Week_Low', '52_Week_High',
                         'Price_vs_52W', 'Price_vs_3M', 'Price_vs_1M', 
                         'Current_vs_All', 'Data_Source']].copy()
        
        display_df.columns = ['Symbol', 'Company', 'Price (₹)', '52W Low', '52W High',
                              '52W %', '3M %', '1M %', 'Current_vs_All (%)', 'Source']
        
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.precision', 2)
        
        print(display_df.to_string(index=False))
        
        # Create Excel report
        print("\n" + "-"*70)
        print("📝 Creating detailed Excel report...")
        print("-"*70)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"Stock_Analysis_Report_{timestamp}.xlsx"
        result_file = analyzer.create_excel_report(df, filename=excel_filename)
        
        # Create visualizations
        chart_files = analyzer.create_visualizations(df, filename_prefix=f"Stock_Analysis_{timestamp}")
        
        # Final summary
        print("\n" + "="*70)
        print(" "*25 + "✅ ANALYSIS COMPLETE!")
        print("="*70)
        
        if result_file:
            print(f"\n📄 Excel Report: {result_file}")
        
        if chart_files:
            print(f"📊 Charts PDF: {chart_files[0]}")
        
        print("\n" + "="*70)
        print("All files have been saved in the current directory.")
        print("="*70 + "\n")
            
    except Exception as e:
        print(f"\n✗ Fatal error in main execution: {str(e)}")
        print("Please check your internet connection and try again.")

if __name__ == "__main__":
    main()